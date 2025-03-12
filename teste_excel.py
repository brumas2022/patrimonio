import pandas as pd
#import openpyxl
from openpyxl import load_workbook

#origem = pd.read_excel("orcamento.xlsx", sheet_name=0)

#print(origem)

#with pd.ExcelWriter('teste_outro.xlsx', mode='a') as writer:  
#    origem.to_excel(writer, sheet_name='Sheet_name_3')
    

w1 = load_workbook("orcamento.xlsx")
w2 = load_workbook("teste_outro.xlsx")

ws_to_move = w1['Planilha1']
ws_to_move._parent = w2
w2._add_sheet(ws_to_move)
w2.save("teste_outro.xlsx")