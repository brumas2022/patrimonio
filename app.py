"""
Sistema de Fiscalização de Contratos Administrativos
SANEAR – CNPJ 03.702.217/0001-31
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import sqlite3
import os
import sys
from datetime import datetime, date

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, HRFlowable, Image as RLImage,
                                     KeepTogether)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# CAMINHOS
# ─────────────────────────────────────────────────────────────────────────────

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DB_PATH   = os.path.join(BASE_DIR, "fiscalizacao.db")

def _logo_path():
    for nome in ("logo_sanear.jpg", "logo_sanear.png"):
        p = os.path.join(BASE_DIR, nome)
        if os.path.exists(p):
            return p
    return None

LOGO_PATH = _logo_path()

CNPJ_SANEAR = "03.702.217/0001-31"
NOME_ORGAO  = "Serviço de Saneamento Ambiental de Rondonópolis Terezinha Silva de Souza"

# ─────────────────────────────────────────────────────────────────────────────
# BANCO DE DADOS
# ─────────────────────────────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS ocorrencias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            num_contrato TEXT NOT NULL,
            mes_ano TEXT NOT NULL,
            tipo TEXT NOT NULL,
            descricao TEXT,
            data_registro TEXT,
            UNIQUE(num_contrato, mes_ano, tipo)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS historico_pdf (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            num_contrato TEXT NOT NULL,
            mes_ano TEXT NOT NULL,
            caminho TEXT,
            data_geracao TEXT
        )
    """)
    # Tabela de aditivos
    c.execute("""
        CREATE TABLE IF NOT EXISTS aditivos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            num_contrato TEXT NOT NULL UNIQUE,
            novo_prazo TEXT,
            novo_valor TEXT,
            observacao TEXT,
            data_registro TEXT
        )
    """)
    conn.commit()
    conn.close()

def salvar_ocorrencia(num_contrato, mes_ano, tipo, descricao):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO ocorrencias (num_contrato, mes_ano, tipo, descricao, data_registro)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(num_contrato, mes_ano, tipo)
        DO UPDATE SET descricao=excluded.descricao, data_registro=excluded.data_registro
    """, (num_contrato, mes_ano, tipo, descricao,
          datetime.now().strftime("%d/%m/%Y %H:%M")))
    conn.commit()
    conn.close()

def carregar_ocorrencia(num_contrato, mes_ano, tipo):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT descricao FROM ocorrencias WHERE num_contrato=? AND mes_ano=? AND tipo=?",
              (num_contrato, mes_ano, tipo))
    row = c.fetchone()
    conn.close()
    return row[0] if row else ""

def salvar_aditivo(num_contrato, novo_prazo, novo_valor, observacao):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO aditivos (num_contrato, novo_prazo, novo_valor, observacao, data_registro)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(num_contrato)
        DO UPDATE SET novo_prazo=excluded.novo_prazo,
                      novo_valor=excluded.novo_valor,
                      observacao=excluded.observacao,
                      data_registro=excluded.data_registro
    """, (num_contrato, novo_prazo, novo_valor, observacao,
          datetime.now().strftime("%d/%m/%Y %H:%M")))
    conn.commit()
    conn.close()

def carregar_aditivo(num_contrato):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT novo_prazo, novo_valor, observacao FROM aditivos WHERE num_contrato=?",
              (num_contrato,))
    row = c.fetchone()
    conn.close()
    return row  # (novo_prazo, novo_valor, observacao) ou None

def registrar_historico_pdf(num_contrato, mes_ano, caminho):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO historico_pdf (num_contrato, mes_ano, caminho, data_geracao) "
              "VALUES (?, ?, ?, ?)",
              (num_contrato, mes_ano, caminho,
               datetime.now().strftime("%d/%m/%Y %H:%M")))
    conn.commit()
    conn.close()

def listar_historico_pdf(num_contrato=None):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if num_contrato:
        c.execute("SELECT num_contrato, mes_ano, caminho, data_geracao FROM historico_pdf "
                  "WHERE num_contrato=? ORDER BY id DESC LIMIT 50", (num_contrato,))
    else:
        c.execute("SELECT num_contrato, mes_ano, caminho, data_geracao FROM historico_pdf "
                  "ORDER BY id DESC LIMIT 100")
    rows = c.fetchall()
    conn.close()
    return rows

def contratos_vencendo(contratos, aditivos_map, dias=30):
    hoje = date.today()
    resultado = []
    for c in contratos:
        ad = aditivos_map.get(c["num_contrato"])
        prazo_str = (ad[0] if ad and ad[0] else None) or c.get("prazo_final", "")
        try:
            dt = datetime.strptime(prazo_str, "%d/%m/%Y").date()
            diff = (dt - hoje).days
            if 0 <= diff <= dias:
                resultado.append((c, diff, prazo_str))
        except Exception:
            pass
    resultado.sort(key=lambda x: x[1])
    return resultado

# ─────────────────────────────────────────────────────────────────────────────
# PLANILHA EXCEL
# ─────────────────────────────────────────────────────────────────────────────

CAMPOS     = ["num_contrato","objeto","contratada","prazo_inicial",
              "prazo_final","valor","fiscal","portaria","data_portaria"]
CABECALHOS = ["Nº Contrato","Objeto","Contratada","Prazo Inicial",
              "Prazo Final","Valor (R$)","Fiscal","Portaria de Nomeação","Data da Portaria"]

def ler_planilha(caminho):
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não instalado.")
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active
    contratos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        c = {}
        for i, campo in enumerate(CAMPOS):
            val = row[i] if i < len(row) else ""
            if val is None:
                val = ""
            if isinstance(val, (datetime, date)):
                val = val.strftime("%d/%m/%Y")
            c[campo] = str(val).strip()
        if c["num_contrato"]:
            contratos.append(c)
    return contratos

# ─────────────────────────────────────────────────────────────────────────────
# GERAÇÃO DO PDF
# ─────────────────────────────────────────────────────────────────────────────

AZUL_SANEAR  = colors.HexColor("#0099D6")
VERDE_SANEAR = colors.HexColor("#5cb85c")
AZUL_ESCURO  = colors.HexColor("#004f82")
CINZA_CAB    = colors.HexColor("#f2f6fa")
CINZA_LINHA  = colors.HexColor("#e4edf5")
CINZA_TEXTO  = colors.HexColor("#2c2c2c")
BRANCO       = colors.white

MESES_NOME = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
               "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

def mes_extenso(mes_ano):
    try:
        m, a = mes_ano.split("/")
        return f"{MESES_NOME[int(m)-1]} de {a}"
    except Exception:
        return mes_ano

def gerar_pdf(contrato, mes_ano, caminho_saida, aditivo=None):
    """
    aditivo = (novo_prazo, novo_valor, observacao) ou None
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab não instalado.")

    PAGE_W, PAGE_H = A4
    MARG_H = 1.8*cm
    MARG_V = 1.6*cm
    larg = PAGE_W - 2*MARG_H          # largura útil ≈ 17.4 cm

    doc = SimpleDocTemplate(
        caminho_saida, pagesize=A4,
        leftMargin=MARG_H, rightMargin=MARG_H,
        topMargin=MARG_V, bottomMargin=MARG_V,
        title="Relatório Mensal de Acompanhamento de Contrato"
    )

    # ── Estilos ───────────────────────────────────────────────────────────────
    def st(name, **kw):
        defaults = dict(fontName="Helvetica", fontSize=9,
                        textColor=CINZA_TEXTO, leading=13)
        defaults.update(kw)
        return ParagraphStyle(name, **defaults)

    s_orgao   = st("orgao",  fontSize=7.5, textColor=colors.HexColor("#555"),
                   leading=11, alignment=TA_LEFT)
    s_titbco  = st("titbco", fontSize=11, fontName="Helvetica-Bold",
                   textColor=BRANCO, alignment=TA_CENTER, leading=15)
    s_ref     = st("ref",    fontSize=9.5, fontName="Helvetica-Bold",
                   textColor=AZUL_ESCURO, alignment=TA_CENTER)
    s_secao   = st("secao",  fontSize=9,  fontName="Helvetica-Bold",
                   textColor=BRANCO, leading=12)
    s_lbl     = st("lbl",    fontSize=7.5, fontName="Helvetica-Bold",
                   textColor=colors.HexColor("#444"), leading=11)
    s_val     = st("val",    fontSize=8.5, leading=12)
    s_corpo   = st("corpo",  fontSize=9,  leading=14, alignment=TA_JUSTIFY)
    s_ass     = st("ass",    fontSize=9,  fontName="Helvetica-Bold",
                   alignment=TA_CENTER)
    s_ass_sub = st("asssub", fontSize=8,  textColor=colors.HexColor("#555"),
                   alignment=TA_CENTER)
    s_aditivo = st("adit",   fontSize=8.5, fontName="Helvetica-Bold",
                   textColor=colors.HexColor("#b05000"), leading=12)

    story = []

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    logo_cell = ""
    if LOGO_PATH and os.path.exists(LOGO_PATH):
        # proporção: 2336×1112 → largura máx 5.5 cm
        logo_w = 5.5*cm
        logo_h = logo_w * (1112/2336)
        logo_cell = RLImage(LOGO_PATH, width=logo_w, height=logo_h)

    orgao_par = Paragraph(
        f"<b>{NOME_ORGAO}</b><br/>"
        f"CNPJ: {CNPJ_SANEAR}  –  Rondonópolis, MT",
        s_orgao)

    cab = Table(
        [[logo_cell, orgao_par]],
        colWidths=[5.7*cm, larg - 5.7*cm],
        style=TableStyle([
            ("VALIGN",      (0,0),(-1,-1),"MIDDLE"),
            ("LEFTPADDING", (0,0),(0,0),   0),
            ("RIGHTPADDING",(0,0),(0,0),  10),
            ("LEFTPADDING", (1,0),(1,0),   8),
        ])
    )
    story.append(cab)
    story.append(Spacer(1, 5))

    # Faixa título azul
    story.append(Table(
        [[Paragraph("RELATÓRIO MENSAL DE ACOMPANHAMENTO DE CONTRATO", s_titbco)]],
        colWidths=[larg],
        style=TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), AZUL_ESCURO),
            ("TOPPADDING",    (0,0),(-1,-1), 7),
            ("BOTTOMPADDING", (0,0),(-1,-1), 7),
        ])
    ))

    # Faixa referência verde-claro
    story.append(Table(
        [[Paragraph(f"Referência: {mes_extenso(mes_ano)}", s_ref)]],
        colWidths=[larg],
        style=TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor("#eaf7ea")),
            ("TOPPADDING",    (0,0),(-1,-1), 4),
            ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ("BOX",           (0,0),(-1,-1), 0.8, VERDE_SANEAR),
        ])
    ))
    story.append(Spacer(1, 8))

    # ── Helpers ───────────────────────────────────────────────────────────────
    def cabsec(num, texto):
        return Table(
            [[Paragraph(f"  {num}. {texto.upper()}", s_secao)]],
            colWidths=[larg],
            style=TableStyle([
                ("BACKGROUND",    (0,0),(-1,-1), AZUL_ESCURO),
                ("TOPPADDING",    (0,0),(-1,-1), 4),
                ("BOTTOMPADDING", (0,0),(-1,-1), 4),
                ("LEFTPADDING",   (0,0),(-1,-1), 6),
                ("LINEBELOW",     (0,0),(-1,-1), 2, VERDE_SANEAR),
            ])
        )

    def tabcampos(pares, ncols=2):
        lw = 3.6*cm
        vw = (larg - lw*ncols) / ncols
        rows = []
        linha = []
        for i,(lb,vl) in enumerate(pares):
            linha.append((lb, vl))
            if len(linha) == ncols or i == len(pares)-1:
                while len(linha) < ncols:
                    linha.append(("",""))
                row = []
                for lb2,vl2 in linha:
                    row.append(Paragraph(lb2, s_lbl))
                    row.append(Paragraph(str(vl2) if vl2 else "—", s_val))
                rows.append(row)
                linha = []
        cw = [lw, vw] * ncols
        est = [
            ("VALIGN",       (0,0),(-1,-1),"TOP"),
            ("TOPPADDING",   (0,0),(-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
            ("LEFTPADDING",  (0,0),(-1,-1), 5),
            ("RIGHTPADDING", (0,0),(-1,-1), 5),
            ("LINEBELOW",    (0,0),(-1,-1), 0.3, colors.HexColor("#d0dce8")),
        ]
        for ri in range(len(rows)):
            bg = CINZA_CAB if ri%2==0 else BRANCO
            est.append(("BACKGROUND",(0,ri),(-1,ri), bg))
        for ci in range(0, ncols*2, 2):
            est.append(("BACKGROUND",(ci,0),(ci,-1), CINZA_LINHA))
        return Table(rows, colWidths=cw, style=TableStyle(est))

    def caixa(texto):
        if not texto:
            texto = "Sem registros para o período."
        return Table(
            [[Paragraph(texto, s_corpo)]],
            colWidths=[larg],
            style=TableStyle([
                ("BOX",           (0,0),(-1,-1), 0.5, colors.HexColor("#a8c0d8")),
                ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor("#f8fbfe")),
                ("TOPPADDING",    (0,0),(-1,-1), 7),
                ("BOTTOMPADDING", (0,0),(-1,-1), 7),
                ("LEFTPADDING",   (0,0),(-1,-1), 9),
                ("RIGHTPADDING",  (0,0),(-1,-1), 9),
            ])
        )

    # ── Seção 1 – Identificação ───────────────────────────────────────────────
    prazo_final_exibir = contrato["prazo_final"]
    valor_exibir       = contrato["valor"]
    nota_aditivo       = ""

    if aditivo:
        np, nv, obs = aditivo
        if np:
            prazo_final_exibir = f"{contrato['prazo_final']} → <b>{np}</b> (Aditivo)"
        if nv:
            valor_exibir = f"{contrato['valor']} → <b>{nv}</b> (Aditivo)"
        if obs:
            nota_aditivo = obs

    pares_id = [
        ("Nº do Contrato:", contrato["num_contrato"]),
        ("Valor (R$):",     valor_exibir),
        ("Objeto:",         contrato["objeto"]),
        ("Contratada:",     contrato["contratada"]),
        ("Prazo Inicial:",  contrato["prazo_inicial"]),
        ("Prazo Final:",    prazo_final_exibir),
    ]

    bloco_id = [
        cabsec(1, "Identificação do Contrato"),
        Spacer(1, 3),
        tabcampos(pares_id, ncols=2),
    ]
    if nota_aditivo:
        bloco_id += [
            Spacer(1, 3),
            Table(
                [[Paragraph(f"⚠  Aditivo registrado: {nota_aditivo}", s_aditivo)]],
                colWidths=[larg],
                style=TableStyle([
                    ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#fff8f0")),
                    ("BOX",(0,0),(-1,-1),0.6,colors.HexColor("#e08020")),
                    ("TOPPADDING",(0,0),(-1,-1),4),
                    ("BOTTOMPADDING",(0,0),(-1,-1),4),
                    ("LEFTPADDING",(0,0),(-1,-1),8),
                ])
            ),
        ]
    bloco_id.append(Spacer(1, 8))
    story.append(KeepTogether(bloco_id))

    # ── Seção 2 – Fiscal ──────────────────────────────────────────────────────
    story.append(KeepTogether([
        cabsec(2, "Dados do Fiscal"),
        Spacer(1, 3),
        tabcampos([
            ("Nome do Fiscal:",       contrato["fiscal"]),
            ("Portaria de Nomeação:", contrato["portaria"]),
            ("Data da Portaria:",     contrato["data_portaria"]),
        ], ncols=2),
        Spacer(1, 8),
    ]))

    # ── Seções 3-6: registros mensais ─────────────────────────────────────────
    secoes = [
        (3,"Ocorrências do Mês",      "ocorrencias"),
        (4,"Diligências Realizadas",  "diligencias"),
        (5,"Avaliação da Execução",   "avaliacoes"),
        (6,"Observações Gerais",      "observacoes"),
    ]
    for n, tit, tipo in secoes:
        texto = carregar_ocorrencia(contrato["num_contrato"], mes_ano, tipo)
        story.append(KeepTogether([
            cabsec(n, tit),
            Spacer(1, 3),
            caixa(texto),
            Spacer(1, 8),
        ]))

    # ── Assinatura (sem seção numerada, compacta) ─────────────────────────────
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor("#b0c4d8")))
    story.append(Spacer(1, 14))

    linha_ass = "_" * 48
    ass_tab = Table([
        [linha_ass],
        [Paragraph(f"<b>{contrato['fiscal']}</b>", s_ass)],
        [Paragraph("Fiscal do Contrato", s_ass_sub)],
        [Paragraph(
            f"Portaria {contrato['portaria']}  –  {contrato['data_portaria']}",
            s_ass_sub)],
        [Paragraph(f"Local e Data: Rondonópolis, ___/___/{datetime.now().year}", s_ass_sub)],
    ],
        colWidths=[larg * 0.52],
        style=TableStyle([
            ("ALIGN",       (0,0),(-1,-1),"CENTER"),
            ("TOPPADDING",  (0,0),(-1,-1), 2),
            ("LEFTPADDING", (0,0),(-1,-1), 0),
        ])
    )
    # Centraliza na página
    story.append(Table([[ass_tab]], colWidths=[larg],
        style=TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),
                           ("VALIGN",(0,0),(-1,-1),"MIDDLE")])))

    doc.build(story)


# ─────────────────────────────────────────────────────────────────────────────
# TEMA VISUAL
# ─────────────────────────────────────────────────────────────────────────────

COR_FUNDO       = "#f0f4f8"
COR_PAINEL      = "#ffffff"
COR_AZUL        = "#0072BB"
COR_AZUL_ESC    = "#004f82"
COR_VERDE       = "#5cb85c"
COR_VERDE_CLARO = "#eaf7ea"
COR_BORDA       = "#c8d8e8"
COR_TEXTO       = "#1a1a1a"
COR_TEXTO_SEC   = "#004f82"
COR_ALERTA      = "#c0392b"
COR_LARANJA     = "#e67e22"
FONTE_TITULO    = ("Segoe UI", 11, "bold")
FONTE_LABEL     = ("Segoe UI", 9)
FONTE_NORMAL    = ("Segoe UI", 9)
FONTE_BOTAO     = ("Segoe UI", 9, "bold")
FONTE_SMALL     = ("Segoe UI", 8)

MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
         "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

TIPOS_REGISTRO = [
    ("ocorrencias","📋 Ocorrências"),
    ("diligencias","🔍 Diligências"),
    ("avaliacoes", "⭐ Avaliação"),
    ("observacoes","📝 Observações"),
]

STATUS_CORES = {
    "Ativo":     "#27ae60",
    "Encerrado": "#7f8c8d",
    "Suspenso":  "#e67e22",
    "Vencendo":  "#c0392b",
}


# ─────────────────────────────────────────────────────────────────────────────
# JANELA DE ADITIVO
# ─────────────────────────────────────────────────────────────────────────────

class JanelaAditivo(tk.Toplevel):
    """Pergunta se há aditivo e coleta novo prazo / novo valor."""

    def __init__(self, parent, num_contrato, prazo_original, callback):
        super().__init__(parent)
        self.title("⚠  Contrato com Prazo Expirado – Registrar Aditivo?")
        self.geometry("480x340")
        self.configure(bg=COR_FUNDO)
        self.resizable(False, False)
        self.grab_set()
        self.callback = callback

        ad_existente = carregar_aditivo(num_contrato)
        np0 = ad_existente[0] if ad_existente and ad_existente[0] else ""
        nv0 = ad_existente[1] if ad_existente and ad_existente[1] else ""
        ob0 = ad_existente[2] if ad_existente and ad_existente[2] else ""

        # Cabeçalho
        tk.Label(self, text="⚠  Contrato com Prazo Expirado",
                 bg="#c0392b", fg="white",
                 font=("Segoe UI", 11, "bold")).pack(fill="x", ipady=8)

        corpo = tk.Frame(self, bg=COR_FUNDO)
        corpo.pack(fill="both", expand=True, padx=18, pady=12)

        tk.Label(corpo,
                 text=f"O prazo original ({prazo_original}) já passou.\n"
                      "Deseja registrar um Termo Aditivo para este contrato?",
                 bg=COR_FUNDO, font=FONTE_LABEL, justify="left",
                 wraplength=430).pack(anchor="w", pady=(0,10))

        # Novo prazo
        f1 = tk.Frame(corpo, bg=COR_FUNDO)
        f1.pack(fill="x", pady=4)
        tk.Label(f1, text="Novo Prazo Final (dd/mm/aaaa):",
                 bg=COR_FUNDO, font=("Segoe UI",9,"bold"), width=28,
                 anchor="w").pack(side="left")
        self.e_prazo = tk.Entry(f1, font=FONTE_NORMAL, width=14,
                                 highlightthickness=1,
                                 highlightbackground=COR_BORDA)
        self.e_prazo.pack(side="left", padx=6)
        self.e_prazo.insert(0, np0)

        # Novo valor
        f2 = tk.Frame(corpo, bg=COR_FUNDO)
        f2.pack(fill="x", pady=4)
        tk.Label(f2, text="Novo Valor (R$) — se alterado:",
                 bg=COR_FUNDO, font=("Segoe UI",9,"bold"), width=28,
                 anchor="w").pack(side="left")
        self.e_valor = tk.Entry(f2, font=FONTE_NORMAL, width=14,
                                 highlightthickness=1,
                                 highlightbackground=COR_BORDA)
        self.e_valor.pack(side="left", padx=6)
        self.e_valor.insert(0, nv0)

        # Observação
        tk.Label(corpo, text="Observação / Nº do Aditivo:",
                 bg=COR_FUNDO, font=("Segoe UI",9,"bold"),
                 anchor="w").pack(fill="x", pady=(8,2))
        self.e_obs = tk.Text(corpo, height=3, font=FONTE_NORMAL,
                              relief="flat", highlightthickness=1,
                              highlightbackground=COR_BORDA, padx=6, pady=4)
        self.e_obs.pack(fill="x")
        if ob0:
            self.e_obs.insert("1.0", ob0)

        # Botões
        bf = tk.Frame(self, bg=COR_FUNDO)
        bf.pack(fill="x", padx=18, pady=10)
        tk.Button(bf, text="✔  Salvar Aditivo", command=self._salvar,
                  bg=COR_AZUL_ESC, fg="white", font=FONTE_BOTAO,
                  relief="flat", padx=10, pady=5, cursor="hand2").pack(side="left")
        tk.Button(bf, text="Ignorar / Sem Aditivo", command=self._ignorar,
                  bg="#888", fg="white", font=FONTE_BOTAO,
                  relief="flat", padx=10, pady=5, cursor="hand2").pack(side="left", padx=8)

    def _salvar(self):
        np = self.e_prazo.get().strip()
        nv = self.e_valor.get().strip()
        ob = self.e_obs.get("1.0","end").strip()
        # Valida data
        if np:
            try:
                datetime.strptime(np, "%d/%m/%Y")
            except ValueError:
                messagebox.showerror("Data inválida",
                                     "Informe o novo prazo no formato dd/mm/aaaa.",
                                     parent=self)
                return
        self.callback(np, nv, ob)
        self.destroy()

    def _ignorar(self):
        self.callback(None, None, None)
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# JANELA HISTÓRICO
# ─────────────────────────────────────────────────────────────────────────────

class JanelaHistorico(tk.Toplevel):
    def __init__(self, parent, num_contrato=None):
        super().__init__(parent)
        self.title("Histórico de Relatórios Gerados")
        self.geometry("700x400")
        self.configure(bg=COR_FUNDO)

        tk.Label(self, text="📄  Histórico de Relatórios PDF",
                 bg=COR_AZUL_ESC, fg="white",
                 font=("Segoe UI",11,"bold")).pack(fill="x", ipady=8)

        cols = ("Contrato","Referência","Gerado em","Arquivo")
        tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        for col in cols:
            tree.heading(col, text=col)
        tree.column("Contrato",   width=120)
        tree.column("Referência", width=100)
        tree.column("Gerado em",  width=130)
        tree.column("Arquivo",    width=330)
        sb = ttk.Scrollbar(self, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True, padx=(10,0), pady=10)
        sb.pack(side="right", fill="y", pady=10, padx=(0,10))

        for row in listar_historico_pdf(num_contrato):
            tree.insert("","end", values=row)

        def abrir(event=None):
            sel = tree.selection()
            if not sel:
                return
            path = tree.item(sel[0])["values"][2]
            if os.path.exists(str(path)):
                _abrir_arquivo(str(path))
            else:
                messagebox.showwarning("Não encontrado",
                    f"Arquivo não encontrado:\n{path}", parent=self)

        tree.bind("<Double-1>", abrir)
        tk.Label(self, text="Duplo clique para abrir o PDF",
                 bg=COR_FUNDO, font=FONTE_SMALL, fg="#666").pack(pady=(0,6))


# ─────────────────────────────────────────────────────────────────────────────
# JANELA ALERTAS
# ─────────────────────────────────────────────────────────────────────────────

class JanelaAlertas(tk.Toplevel):
    def __init__(self, parent, alertas):
        super().__init__(parent)
        self.title("⚠ Contratos Vencendo em Breve")
        self.geometry("620x360")
        self.configure(bg=COR_FUNDO)

        tk.Label(self, text="⚠  Contratos com prazo se encerrando (próximos 30 dias)",
                 bg="#c0392b", fg="white",
                 font=("Segoe UI",11,"bold")).pack(fill="x", ipady=8)

        cols = ("Nº Contrato","Contratada","Prazo Vigente","Dias Restantes")
        tree = ttk.Treeview(self, columns=cols, show="headings", height=14)
        for col in cols:
            tree.heading(col, text=col)
        tree.column("Nº Contrato",    width=120)
        tree.column("Contratada",     width=210)
        tree.column("Prazo Vigente",  width=110)
        tree.column("Dias Restantes", width=110)
        sb = ttk.Scrollbar(self, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True, padx=(10,0), pady=10)
        sb.pack(side="right", fill="y", pady=10, padx=(0,10))

        for c, dias, prazo in alertas:
            tag = "red" if dias <= 7 else "orange"
            tree.insert("","end",
                        values=(c["num_contrato"], c["contratada"],
                                prazo, f"{dias} dia(s)"),
                        tags=(tag,))
        tree.tag_configure("red",    foreground="#c0392b")
        tree.tag_configure("orange", foreground="#e67e22")


# ─────────────────────────────────────────────────────────────────────────────
# UTILITÁRIO
# ─────────────────────────────────────────────────────────────────────────────

def _abrir_arquivo(caminho):
    if sys.platform == "win32":
        os.startfile(caminho)
    elif sys.platform == "darwin":
        os.system(f'open "{caminho}"')
    else:
        os.system(f'xdg-open "{caminho}"')


# ─────────────────────────────────────────────────────────────────────────────
# APLICAÇÃO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        init_db()
        self.title("Fiscalização de Contratos – SANEAR")
        self.geometry("1260x820")
        self.minsize(1000,680)
        self.configure(bg=COR_FUNDO)

        self.contratos          = []
        self.contrato_selecionado = None
        self.status_contratos   = {}    # num -> status
        self.aditivos_map       = {}    # num -> (novo_prazo, novo_valor, obs)

        self._build_ui()
        self._criar_planilha_modelo()
        self._carregar_aditivos_db()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _carregar_aditivos_db(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT num_contrato, novo_prazo, novo_valor, observacao FROM aditivos")
        for row in c.fetchall():
            self.aditivos_map[row[0]] = (row[1], row[2], row[3])
        conn.close()

    def _btn(self, parent, texto, cmd, cor, pad=8):
        b = tk.Button(parent, text=texto, command=cmd,
                      bg=cor, fg="white", font=FONTE_BOTAO,
                      relief="flat", bd=0, cursor="hand2",
                      padx=pad, pady=5,
                      activebackground=self._esc(cor),
                      activeforeground="white")
        b.bind("<Enter>", lambda e, c=cor: b.config(bg=self._esc(c)))
        b.bind("<Leave>", lambda e, c=cor: b.config(bg=c))
        return b

    @staticmethod
    def _esc(hx):
        try:
            r=max(0,int(hx[1:3],16)-25)
            g=max(0,int(hx[3:5],16)-25)
            b=max(0,int(hx[5:7],16)-25)
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception:
            return hx

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        # Barra superior
        topo = tk.Frame(self, bg=COR_AZUL_ESC, height=62)
        topo.pack(fill="x")
        topo.pack_propagate(False)

        if LOGO_PATH:
            try:
                from PIL import Image, ImageTk
                img = Image.open(LOGO_PATH)
                iw, ih = 148, int(148*(1112/2336))
                img = img.resize((iw, ih), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                tk.Label(topo, image=self._logo_img,
                         bg=COR_AZUL_ESC).pack(side="left", padx=14, pady=8)
            except Exception:
                tk.Label(topo, text="SANEAR", bg=COR_AZUL_ESC, fg="white",
                         font=("Segoe UI",15,"bold")).pack(side="left", padx=14)
        else:
            tk.Label(topo, text="SANEAR", bg=COR_AZUL_ESC, fg="white",
                     font=("Segoe UI",15,"bold")).pack(side="left", padx=14)

        tk.Label(topo,
                 text="Sistema de Fiscalização de Contratos Administrativos",
                 bg=COR_AZUL_ESC, fg="#cce4f5",
                 font=("Segoe UI",10)).pack(side="left", padx=6)

        self._btn(topo,"⚠ Vencimentos",    self._ver_alertas,  "#c0392b",10).pack(side="right",padx=8,pady=12)
        self._btn(topo,"📄 Histórico PDF",  self._ver_historico, COR_VERDE,10).pack(side="right",padx=4,pady=12)

        # Corpo
        corpo = tk.Frame(self, bg=COR_FUNDO)
        corpo.pack(fill="both", expand=True, padx=12, pady=10)

        # ── Painel esquerdo ───────────────────────────────────────────────────
        esq = tk.Frame(corpo, bg=COR_PAINEL,
                        highlightthickness=1, highlightbackground=COR_BORDA)
        esq.pack(side="left", fill="y", padx=(0,10))
        esq.pack_propagate(False)
        esq.configure(width=282)

        tk.Frame(esq, bg=COR_AZUL_ESC, height=36).pack(fill="x")
        tk.Label(esq, text="  Contratos Importados", bg=COR_AZUL_ESC, fg="white",
                 font=FONTE_TITULO).place(x=0, y=0, width=282, height=36)

        bf = tk.Frame(esq, bg=COR_PAINEL)
        bf.pack(fill="x", padx=8, pady=(44,0))
        self._btn(bf,"📂  Importar Planilha",  self._importar,     COR_AZUL, 6).pack(fill="x",pady=2)
        self._btn(bf,"📋  Modelo Excel",        self._abrir_modelo, "#5a8f6a",6).pack(fill="x",pady=2)

        # Busca
        sb_frame = tk.Frame(esq, bg=COR_PAINEL)
        sb_frame.pack(fill="x", padx=8, pady=(6,4))
        tk.Label(sb_frame, text="🔎", bg=COR_PAINEL, font=FONTE_SMALL).pack(side="left")
        self.var_busca = tk.StringVar()
        self.var_busca.trace("w", self._filtrar_lista)
        tk.Entry(sb_frame, textvariable=self.var_busca, font=FONTE_SMALL,
                 relief="flat", highlightthickness=1,
                 highlightbackground=COR_BORDA).pack(side="left",fill="x",expand=True,padx=(4,0))

        # Lista
        lf = tk.Frame(esq, bg=COR_PAINEL)
        lf.pack(fill="both", expand=True, padx=6, pady=(0,6))
        scv = ttk.Scrollbar(lf)
        scv.pack(side="right", fill="y")
        self.lista = tk.Listbox(lf, yscrollcommand=scv.set,
                                 font=FONTE_NORMAL, bg="white", fg=COR_TEXTO,
                                 relief="flat", selectbackground=COR_AZUL,
                                 selectforeground="white", activestyle="none",
                                 borderwidth=0,
                                 highlightthickness=1, highlightbackground=COR_BORDA)
        self.lista.pack(side="left", fill="both", expand=True)
        scv.config(command=self.lista.yview)
        self.lista.bind("<<ListboxSelect>>", self._selecionar_contrato)

        self.var_contador = tk.StringVar(value="0 contratos")
        tk.Label(esq, textvariable=self.var_contador,
                 bg=COR_PAINEL, font=FONTE_SMALL, fg="#888").pack(pady=(0,6))

        # ── Painel direito ────────────────────────────────────────────────────
        dir_frame = tk.Frame(corpo, bg=COR_FUNDO)
        dir_frame.pack(side="left", fill="both", expand=True)

        # Dados do contrato
        self.painel_dados = tk.LabelFrame(dir_frame, text=" Dados do Contrato Selecionado ",
                                           bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                                           font=FONTE_TITULO, relief="flat",
                                           highlightthickness=1, highlightbackground=COR_BORDA)
        self.painel_dados.pack(fill="x", pady=(0,8))
        self.labels_dados = {}
        self._build_painel_dados()

        # Banner de aditivo
        self.frame_aditivo = tk.Frame(dir_frame, bg="#fff3e0",
                                       highlightthickness=1,
                                       highlightbackground="#e08020")
        self.lbl_aditivo = tk.Label(self.frame_aditivo, bg="#fff3e0",
                                     font=("Segoe UI",8,"bold"),
                                     fg="#b05000", anchor="w")
        self.lbl_aditivo.pack(side="left", padx=10, pady=4)
        self._btn(self.frame_aditivo, "✏ Editar Aditivo",
                  self._editar_aditivo, "#e08020", 8).pack(side="right", padx=8, pady=4)
        # Não exibe até ter um contrato com aditivo

        # Controles mês/ano/status
        ctrl = tk.Frame(dir_frame, bg=COR_PAINEL,
                         highlightthickness=1, highlightbackground=COR_BORDA)
        ctrl.pack(fill="x", pady=(0,8))

        tk.Label(ctrl, text="Mês:", bg=COR_PAINEL, font=FONTE_LABEL).pack(side="left",padx=(12,4),pady=8)
        self.var_mes = tk.StringVar(value=MESES[datetime.now().month-1])
        ttk.Combobox(ctrl, textvariable=self.var_mes, values=MESES,
                     width=12, state="readonly").pack(side="left", padx=2)

        tk.Label(ctrl, text="Ano:", bg=COR_PAINEL, font=FONTE_LABEL).pack(side="left",padx=(8,4))
        self.var_ano = tk.StringVar(value=str(datetime.now().year))
        anos = [str(a) for a in range(2020, datetime.now().year+3)]
        ttk.Combobox(ctrl, textvariable=self.var_ano, values=anos,
                     width=7, state="readonly").pack(side="left", padx=2)

        self._btn(ctrl,"🔍 Carregar", self._carregar_registros, COR_AZUL).pack(side="left",padx=12)

        tk.Label(ctrl, text="Status:", bg=COR_PAINEL, font=FONTE_LABEL).pack(side="left",padx=(20,4))
        self.var_status = tk.StringVar(value="Ativo")
        ttk.Combobox(ctrl, textvariable=self.var_status,
                     values=list(STATUS_CORES.keys()),
                     width=12, state="readonly").pack(side="left", padx=2)
        self.var_status.trace("w", self._mudar_status)
        self.lbl_status_cor = tk.Label(ctrl, text="●", bg=COR_PAINEL,
                                        font=("Segoe UI",14),
                                        fg=STATUS_CORES["Ativo"])
        self.lbl_status_cor.pack(side="left", padx=4)

        # Notebook
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook", background=COR_FUNDO, borderwidth=0)
        style.configure("TNotebook.Tab", font=FONTE_LABEL, padding=[12,6],
                         background="#dde8f0", foreground=COR_TEXTO)
        style.map("TNotebook.Tab",
                  background=[("selected", COR_AZUL_ESC)],
                  foreground=[("selected","white")])

        self.notebook = ttk.Notebook(dir_frame)
        self.notebook.pack(fill="both", expand=True)

        self.areas_texto = {}
        for tipo, titulo in TIPOS_REGISTRO:
            frm = tk.Frame(self.notebook, bg=COR_PAINEL)
            self.notebook.add(frm, text=f" {titulo} ")
            tk.Label(frm, text=titulo.split(" ",1)[-1],
                     bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                     font=FONTE_TITULO).pack(anchor="w", padx=12, pady=(8,2))
            tf = tk.Frame(frm, bg=COR_PAINEL)
            tf.pack(fill="both", expand=True, padx=12, pady=(0,8))
            sbv = tk.Scrollbar(tf)
            sbv.pack(side="right", fill="y")
            txt = tk.Text(tf, yscrollcommand=sbv.set,
                          font=("Segoe UI",10), relief="flat",
                          bg="#f7fafd", fg=COR_TEXTO, wrap="word",
                          padx=10, pady=8,
                          highlightthickness=1, highlightbackground=COR_BORDA)
            txt.pack(fill="both", expand=True)
            sbv.config(command=txt.yview)
            self.areas_texto[tipo] = txt

        # Barra de ações
        barra = tk.Frame(dir_frame, bg=COR_PAINEL,
                          highlightthickness=1, highlightbackground=COR_BORDA)
        barra.pack(fill="x", pady=(8,0))
        self._btn(barra,"💾  Salvar Registros",    self._salvar_registros, COR_AZUL_ESC,12).pack(side="left",padx=10,pady=8)
        self._btn(barra,"📄  Gerar PDF do Relatório",self._gerar_pdf,       COR_LARANJA, 12).pack(side="left",padx=4, pady=8)
        self._btn(barra,"📂  Histórico de PDFs",   self._ver_historico,    "#7f8c8d",   10).pack(side="left",padx=4, pady=8)

        self.status_var = tk.StringVar(value="Importe uma planilha Excel para começar.")
        tk.Label(barra, textvariable=self.status_var,
                 bg=COR_PAINEL, font=FONTE_SMALL, fg="#555").pack(side="right",padx=12)

    def _build_painel_dados(self):
        campos = [
            ("num_contrato", "Nº Contrato",   0,0),
            ("objeto",       "Objeto",         0,2),
            ("contratada",   "Contratada",     1,0),
            ("valor",        "Valor (R$)",     1,2),
            ("prazo_inicial","Prazo Inicial",  2,0),
            ("prazo_final",  "Prazo Final",    2,2),
            ("fiscal",       "Fiscal",         3,0),
            ("portaria",     "Portaria",       3,2),
            ("data_portaria","Data Portaria",  4,0),
        ]
        for campo, label, row, col in campos:
            tk.Label(self.painel_dados, text=label+":",
                     bg=COR_PAINEL, font=("Segoe UI",8,"bold"),
                     fg="#555").grid(row=row, column=col, sticky="w", padx=(14,2), pady=3)
            var = tk.StringVar(value="—")
            tk.Label(self.painel_dados, textvariable=var,
                     bg="#eaf3fb", font=FONTE_NORMAL, fg=COR_TEXTO,
                     relief="flat", padx=6, pady=2,
                     width=26, anchor="w").grid(row=row, column=col+1,
                                                 sticky="ew", padx=(0,12), pady=3)
            self.labels_dados[campo] = var
        for i in [1,3,5,7]:
            self.painel_dados.grid_columnconfigure(i, weight=1)

    # ── Ações ─────────────────────────────────────────────────────────────────

    def _importar(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Erro","Instale:\npip install openpyxl"); return
        caminho = filedialog.askopenfilename(
            title="Selecionar Planilha",
            filetypes=[("Excel","*.xlsx *.xls"),("Todos","*.*")])
        if not caminho:
            return
        try:
            self.contratos = ler_planilha(caminho)
            self._atualizar_lista(self.contratos)
            alertas = contratos_vencendo(self.contratos, self.aditivos_map, 30)
            msg = f"✔ {len(self.contratos)} contrato(s) importado(s)."
            if alertas:
                msg += f"  ⚠ {len(alertas)} vencendo em 30 dias."
            self.status_var.set(msg)
        except Exception as e:
            messagebox.showerror("Erro ao importar", str(e))

    def _atualizar_lista(self, lista):
        self.lista.delete(0,"end")
        icones = {"Ativo":"🟢","Encerrado":"⚫","Suspenso":"🟡","Vencendo":"🔴"}
        for c in lista:
            st = self.status_contratos.get(c["num_contrato"],"Ativo")
            ic = icones.get(st,"🟢")
            ad = self.aditivos_map.get(c["num_contrato"])
            sufixo = " ✎" if ad and ad[0] else ""
            self.lista.insert("end", f" {ic}  {c['num_contrato']}{sufixo}")
        self.var_contador.set(f"{len(lista)} contrato(s)")

    def _filtrar_lista(self, *args):
        t = self.var_busca.get().lower()
        fl = [c for c in self.contratos
              if t in c["num_contrato"].lower()
              or t in c["contratada"].lower()
              or t in c["objeto"].lower()]
        self._atualizar_lista(fl)

    def _get_filtrados(self):
        t = self.var_busca.get().lower()
        if t:
            return [c for c in self.contratos
                    if t in c["num_contrato"].lower()
                    or t in c["contratada"].lower()
                    or t in c["objeto"].lower()]
        return self.contratos

    def _selecionar_contrato(self, event=None):
        sel = self.lista.curselection()
        if not sel:
            return
        fl = self._get_filtrados()
        idx = sel[0]
        self.contrato_selecionado = fl[idx] if idx < len(fl) else None
        if not self.contrato_selecionado:
            return
        for campo, var in self.labels_dados.items():
            var.set(self.contrato_selecionado.get(campo,"—") or "—")
        st = self.status_contratos.get(self.contrato_selecionado["num_contrato"],"Ativo")
        self.var_status.set(st)
        self._atualizar_banner_aditivo()
        self._carregar_registros()
        self._verificar_prazo_expirado()

    def _atualizar_banner_aditivo(self):
        if not self.contrato_selecionado:
            self.frame_aditivo.pack_forget()
            return
        ad = self.aditivos_map.get(self.contrato_selecionado["num_contrato"])
        if ad and (ad[0] or ad[1]):
            partes = []
            if ad[0]: partes.append(f"Novo Prazo: {ad[0]}")
            if ad[1]: partes.append(f"Novo Valor: R$ {ad[1]}")
            if ad[2]: partes.append(f"Obs: {ad[2]}")
            self.lbl_aditivo.config(text="⚠  ADITIVO REGISTRADO  –  " + "  |  ".join(partes))
            self.frame_aditivo.pack(fill="x", pady=(0,6), before=self.notebook)
        else:
            self.frame_aditivo.pack_forget()

    def _verificar_prazo_expirado(self):
        """Se prazo final (considerando aditivo) já passou, oferece registro de aditivo."""
        if not self.contrato_selecionado:
            return
        num = self.contrato_selecionado["num_contrato"]
        ad = self.aditivos_map.get(num)
        prazo_str = (ad[0] if ad and ad[0] else None) or \
                    self.contrato_selecionado.get("prazo_final","")
        try:
            dt = datetime.strptime(prazo_str, "%d/%m/%Y").date()
            if dt < date.today():
                def salvar_cb(np, nv, ob):
                    if np is not None or nv is not None:
                        salvar_aditivo(num, np or "", nv or "", ob or "")
                        self.aditivos_map[num] = (np or "", nv or "", ob or "")
                        self.labels_dados["prazo_final"].set(
                            f"{self.contrato_selecionado['prazo_final']} → {np}" if np
                            else self.contrato_selecionado["prazo_final"])
                        self._atualizar_banner_aditivo()
                        self._atualizar_lista(self.contratos)
                JanelaAditivo(self, num,
                              self.contrato_selecionado["prazo_final"],
                              salvar_cb)
        except Exception:
            pass

    def _editar_aditivo(self):
        if not self.contrato_selecionado:
            return
        num = self.contrato_selecionado["num_contrato"]
        prazo_orig = self.contrato_selecionado.get("prazo_final","")
        def cb(np, nv, ob):
            if np is not None:
                salvar_aditivo(num, np or "", nv or "", ob or "")
                self.aditivos_map[num] = (np or "", nv or "", ob or "")
                self._atualizar_banner_aditivo()
                self._atualizar_lista(self.contratos)
        JanelaAditivo(self, num, prazo_orig, cb)

    def _carregar_registros(self):
        if not self.contrato_selecionado:
            return
        num = self.contrato_selecionado["num_contrato"]
        mes_ano = self._mes_ano_atual()
        for tipo,_ in TIPOS_REGISTRO:
            txt = self.areas_texto[tipo]
            txt.delete("1.0","end")
            c = carregar_ocorrencia(num, mes_ano, tipo)
            if c:
                txt.insert("1.0", c)
        self.status_var.set(f"📂 Registros de {mes_ano} – Contrato {num}")

    def _salvar_registros(self):
        if not self.contrato_selecionado:
            messagebox.showwarning("Atenção","Selecione um contrato primeiro."); return
        num = self.contrato_selecionado["num_contrato"]
        mes_ano = self._mes_ano_atual()
        for tipo,_ in TIPOS_REGISTRO:
            texto = self.areas_texto[tipo].get("1.0","end").strip()
            salvar_ocorrencia(num, mes_ano, tipo, texto)
        self.status_var.set(f"✔ Registros de {mes_ano} salvos!")
        messagebox.showinfo("Salvo", f"Registros de {mes_ano} salvos com sucesso!")

    def _gerar_pdf(self):
        if not self.contrato_selecionado:
            messagebox.showwarning("Atenção","Selecione um contrato primeiro."); return
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("Erro","Instale:\npip install reportlab"); return
        self._salvar_registros()
        mes_ano = self._mes_ano_atual()
        num_safe = self.contrato_selecionado["num_contrato"].replace("/","-")
        nome_sug = f"Relatorio_{num_safe}_{mes_ano.replace('/','_')}.pdf"
        caminho = filedialog.asksaveasfilename(
            title="Salvar Relatório PDF",
            defaultextension=".pdf", initialfile=nome_sug,
            filetypes=[("PDF","*.pdf")])
        if not caminho:
            return
        try:
            ad = self.aditivos_map.get(self.contrato_selecionado["num_contrato"])
            gerar_pdf(self.contrato_selecionado, mes_ano, caminho, aditivo=ad)
            registrar_historico_pdf(
                self.contrato_selecionado["num_contrato"], mes_ano, caminho)
            self.status_var.set(f"✔ PDF gerado: {os.path.basename(caminho)}")
            messagebox.showinfo("PDF Gerado", f"Relatório salvo em:\n{caminho}")
            _abrir_arquivo(caminho)
        except Exception as e:
            messagebox.showerror("Erro ao gerar PDF", str(e))

    def _mudar_status(self, *args):
        if not self.contrato_selecionado:
            return
        novo = self.var_status.get()
        self.status_contratos[self.contrato_selecionado["num_contrato"]] = novo
        self.lbl_status_cor.config(fg=STATUS_CORES.get(novo, COR_AZUL))
        self._atualizar_lista(self.contratos)

    def _ver_historico(self):
        num = self.contrato_selecionado["num_contrato"] if self.contrato_selecionado else None
        JanelaHistorico(self, num)

    def _ver_alertas(self):
        if not self.contratos:
            messagebox.showinfo("Alertas","Importe uma planilha para verificar vencimentos."); return
        alertas = contratos_vencendo(self.contratos, self.aditivos_map, 30)
        if not alertas:
            messagebox.showinfo("Alertas","✅ Nenhum contrato vencendo nos próximos 30 dias.")
        else:
            JanelaAlertas(self, alertas)

    def _mes_ano_atual(self):
        mes = MESES.index(self.var_mes.get()) + 1
        return f"{mes:02d}/{self.var_ano.get()}"

    def _criar_planilha_modelo(self):
        if not OPENPYXL_AVAILABLE:
            return
        caminho = os.path.join(BASE_DIR, "modelo_contratos.xlsx")
        if os.path.exists(caminho):
            return
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contratos"
            fill_cab = PatternFill("solid", fgColor="004f82")
            font_cab = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
            borda = Border(
                left=Side(style="thin",color="AAAAAA"),
                right=Side(style="thin",color="AAAAAA"),
                top=Side(style="thin",color="AAAAAA"),
                bottom=Side(style="thin",color="AAAAAA"))
            for col, cab in enumerate(CABECALHOS,1):
                cell = ws.cell(row=1, column=col, value=cab)
                cell.font = font_cab; cell.fill = fill_cab
                cell.alignment = Alignment(horizontal="center",vertical="center")
                cell.border = borda
            ws.row_dimensions[1].height = 22
            larguras = [18,35,30,14,14,16,25,22,16]
            for col,larg in enumerate(larguras,1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = larg
            exemplo = ["2024/001","Prestação de serviços de limpeza e conservação",
                       "Empresa Exemplo Ltda.","01/01/2024","31/12/2024",
                       "120.000,00","João da Silva","Portaria nº 001/2024","02/01/2024"]
            fill_ex = PatternFill("solid", fgColor="eaf3fb")
            for col,val in enumerate(exemplo,1):
                cell = ws.cell(row=2,column=col,value=val)
                cell.fill = fill_ex; cell.border = borda
                cell.font = Font(name="Segoe UI",size=9)
            wb.save(caminho)
        except Exception:
            pass

    def _abrir_modelo(self):
        caminho = os.path.join(BASE_DIR, "modelo_contratos.xlsx")
        if not os.path.exists(caminho):
            self._criar_planilha_modelo()
        if os.path.exists(caminho):
            _abrir_arquivo(caminho)


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
