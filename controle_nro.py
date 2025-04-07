import streamlit as st 
import pandas as pd
from openpyxl import load_workbook

def criar_nro():
    w1 = load_workbook("nova.xlsx")
    aba = w1.active
    aba.append([nome, assunto])
    #aba.max_row
    number=aba.max_row-2
    w1.save("nova.xlsx")
    df = pd.read_excel("nova.xlsx", sheet_name=0)
    st.write("o numero do seu memorando é:", number)
    st.dataframe(df)
    
    pass

with st.form("Numeros"):
    lista = ["Geral", "Tecnica", "Administrativo", "Manutenção", "Juridico"]
    
    choice = st.sidebar.selectbox("Escolha o setor", lista)
    
    nome = st.text_input("Qual o seu nome?")
    
    assunto = st.text_input("Relate o assunto do memorando")
    
    enviar = st.form_submit_button("CONFIRME")
    
    if enviar:
        if choice==lista[0]:
           criar_nro() 
           
        elif choice==lista[1]:
            criar_nro(1)
            
        elif choice==lista[2]:
            criar_nro(2)
            
