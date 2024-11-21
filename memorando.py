import streamlit as st 
import pandas as pd
from openpyxl import load_workbook


st.set_page_config("Solicitação de materiais e orçamentos", layout="wide")
st.image("logosanear.png", width=300)
def repete():
    item = int(input("Qtos itens teremos? "))
    itens = 0
    lista_nome = []
    while item > itens:
      nome = input("Digite seu nome: ")
      #sobrenome = input("Digite sobrenome: ")
      #idade = input("Digite idade: ")
      lista_nome.append(nome)
      item = item - 1
    print(lista_nome)

def criar_memo():
    memo = load_workbook("modelo_memo.xlsx")
    aba_modelo = memo["Plan1"]
    #nro=6
    aba_modelo.cell(row=3, column=5).value = f"n° {nro_memo}"
    
    #while nro > 0:
    #      print("Exeemplo", nro)  
    #      nro = nro - 1
          
   
   
    aba_modelo.cell(row=22, column=1).value = "1"
    aba_modelo.cell(row=23, column=1).value = "2"
    aba_modelo.cell(row=24, column=1).value = "3"
    aba_modelo.cell(row=25, column=1).value = "4"
    aba_modelo.cell(row=26, column=1).value = "5"
    aba_modelo.cell(row=27, column=1).value = "6"
    aba_modelo.cell(row=28, column=1).value = "7"
    
    aba_modelo.cell(row=22, column=2).value = lista_orc[0]
    aba_modelo.cell(row=23, column=2).value = lista_orc[1]
    aba_modelo.cell(row=24, column=2).value = lista_orc[2]
    aba_modelo.cell(row=25, column=2).value = lista_orc[3]
    aba_modelo.cell(row=26, column=2).value = lista_orc[4]
    aba_modelo.cell(row=27, column=2).value = lista_orc[5]

    aba_modelo.cell(row=22, column=4).value = descr_orc[0]
    aba_modelo.cell(row=23, column=4).value = descr_orc[1]
    aba_modelo.cell(row=24, column=4).value = descr_orc[2]
    aba_modelo.cell(row=25, column=4).value = descr_orc[3]
    aba_modelo.cell(row=26, column=4).value = descr_orc[4]
    aba_modelo.cell(row=27, column=4).value = descr_orc[5]
    
    aba_modelo.cell(row=22, column=7).value = qtdes[0]
    aba_modelo.cell(row=23, column=7).value = qtdes[1]
    aba_modelo.cell(row=24, column=7).value = qtdes[2]
    aba_modelo.cell(row=25, column=7).value = qtdes[3]
    aba_modelo.cell(row=26, column=7).value = qtdes[4]
    aba_modelo.cell(row=27, column=7).value = qtdes[5]
    
    
    
    
    #print(aba_alunos.cell(row=1, column=2).value)
    memo.save(f"C:/Users/Compras/Desktop/2024/Almoxarifado e Patrimonio/{nome_arquivo}.xlsx")
    st.warning("O arquivo foi salvo como")

    

def orcamento():
    
   global lista_orc
   global descr_orc
   global qtdes
   global nro_itens
    
   nro_itens = st.number_input("Digite a quantidade de itens a serem solicitados - funciona com 6", value=0)
   item_zero = 0
   lista_orc=[]
   descr_orc=[]
   qtdes=[]
   nro_qtde = 50
   nro_descr = 100
   
   df=pd.read_excel("Estoque_Data_Atual_Excel.xlsx", sheet_name=0)
   df.columns=['Item', 'Descricao', 'Unidade', 'Qtde', 'ValorUnit', 'ValorTotal']
   nomes_orc = df['Item'].tolist()
     
   
   col=st.columns([0.25,0.5,0.25])
   col[0].text("Numero do produto")
   col[1].text("Descrição")
   col[2].text("Quantidade")
  
   # INSERIR OS PRODUTOS DO ESTOQUE PARA AQUISICAO
   
   global item_orc1, item_orc2, item_orc3, item_orc4, item_orc5, item_orc6
   global qtde1, qtde2, qtde3, qtde4, qtde5, qtde6
   global descricao1, descricao2, descricao3, descricao4, descricao5, descricao6
   
   while nro_itens > item_zero:
          item_orc=col[0].selectbox("Numero", nomes_orc, key=nro_itens, label_visibility="collapsed")
          resultado_item=df[df['Item']==item_orc]
          descricao=resultado_item.iat[0,1]
          col[1].text_input("Descrição", descricao, key=nro_descr, label_visibility="collapsed")
          nro_qtde = nro_qtde + 1 
          nro_descr= nro_descr + 1
          qtde=col[2].text_input("Quantidade", key=nro_qtde, label_visibility="collapsed")
          lista_orc.append(item_orc)
          descr_orc.append(descricao)
          qtdes.append(qtde)
          nro_itens = nro_itens-1
          
 
   
         
   #with col[0]:
   #    item_orc1=col[0].selectbox("Numero do produto", nomes_orc)
   #    resultado_item1 = df[df['Item']==item_orc1]
       
   #    descricao1=resultado_item1.iat[0,1]
   #col[1].write("Descrição do produto")
   #col[1].write(descricao1)
  
   #qtde1=col[2].text_input("Qtde produto 1 : ")
   
   
   
   #item_orc2=col[0].selectbox("Numero do produto :", nomes_orc)
   #resultado_item2 = df[df['Item']==item_orc2]
   #descricao2=resultado_item2.iat[0,1]
   #col[1].write("Descrição do produto")
   #col[1].write(descricao2)
   #qtde2=col[2].text_input("Qtde produto 2 : ")
   
   
   #item_orc3=col[0].selectbox("Descrição 3 :", nomes_orc)
   #resultado_item3 = df[df['Item']==item_orc3]
   #descricao3=resultado_item3.iat[0,1]
   #col[1].write("Descrição do produto")
   #col[1].write(descricao3)
   #qtde3=col[2].text_input("Qtde produto 3 : ")

   #item_orc4=col[0].selectbox("Descrição 4 :", nomes_orc)
   #resultado_item4 = df[df['Item']==item_orc4]
   #descricao4=resultado_item4.iat[0,1]
   #col[1].write("Descrição do produto")
   #col[1].write(descricao4)
   #qtde4=col[2].text_input("Qtde produto 4 : ")
   
   #item_orc5=col[0].selectbox("Descrição 5 :", nomes_orc)
   #resultado_item5 = df[df['Item']==item_orc5]
   #descricao5=resultado_item5.iat[0,1]
   #col[1].write("Descrição do produto")
   #col[1].write(descricao5)
   #qtde5=col[2].text_input("Qtde produto 5 : ")
   
   #item_orc6=col[0].selectbox("Descrição 6 :", nomes_orc)
   #resultado_item6 = df[df['Item']==item_orc6]
   #descricao6=resultado_item6.iat[0,1]
   #col[1].write("Descrição do produto")
   #col[1].write(descricao6)
   #qtde6=col[2].text_input("Qtde produto 6 : ")
   
   # CRIAR AS LISTAS 
   #lista_orc.append(item_orc1)
   #lista_orc.append(item_orc2)
   #lista_orc.append(item_orc3)
   #lista_orc.append(item_orc4)
   #lista_orc.append(item_orc5)
   #lista_orc.append(item_orc6)
   #descr_orc.append(descricao1)
   #descr_orc.append(descricao2)
   #descr_orc.append(descricao3)
   #descr_orc.append(descricao4)
   #descr_orc.append(descricao5)
   #descr_orc.append(descricao6)
   #qtdes.append(qtde1)
   #qtdes.append(qtde2)
   #qtdes.append(qtde3)
   #qtdes.append(qtde4)
   #qtdes.append(qtde5)
   #qtdes.append(qtde6)

   # MOSTRAR NA TELA OS ITENS ESCOLHIDOS
   lista_total=list(zip(lista_orc, descr_orc, qtdes))
   df_orc=pd.DataFrame(lista_total, columns=['Numero', 'Descricao', 'Qtde'])
                       
   st.dataframe(df_orc)
   #df_orc.to_excel("orcamento.xlsx", index=False)
   #col[1].dataframe(qtdes)
   # df_orc.to_excel("orcamento.xlsx", sheet_name='Planilha1') ##criar tabela em excel com os dados do dataframe criado

   

   # DISPARAR PARA FORNECEDORES
   #enviar = col[0].button("ENVIAR POR EMAIL PARA FORNECEDORES", on_click=imprimir)
   #enviar = col[0].button("ENVIAR POR EMAIL PARA FORNECEDORES")
   #col[0].link_button("ENVIAR ORCAMENTO", "https://mail.google.com/mail/u/0/#inbox?compose=new")
   #col[0].link_button("ENVIAR ORCAMENTO", "https://mail.google.com/mail/u/0/#inbox?compose=GTvVlcSKkHdPzwgWkcrNGPSHrrKdTGvlGShJwtqqrtkpppgzfSbRtqKjbKXrwzFCvbNDLrMzCrWhP")
   #from openpyxl.workbook import Workbook
   #writer = pd.ExcelWriter("/mount/src/patrimonio/amostra.xlsx")
   #df_orc.to_excel(writer, sheet_name='Planilha1')
   #writer._save()

orcamento()
global nome_arquivo, nro_memo
nome_arquivo=st.text_input("Qual o nome do arquivo? ")
nro_memo = st.text_input("Digite o nro do memorando : ")
criar_orcamento=enviar = st.button("CRIAR MEMORANDO")
if criar_orcamento:
    
    criar_memo()
