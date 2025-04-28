import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook

def criar(nome, ind):
    wb = openpyxl.Workbook()
    del wb['Sheet']
    if ind==0:
        wb.create_sheet(title=f'{nome}', index=ind)
        wb.save("nome_teste.xlsx")
    else:
        pass   
        
    wb.save("nome1.xlsx")
    
    
def botao():
    st.markdown('<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.5/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-SgOJa3DmI69IUzQ2PVdRZhwQ+dy64/BUtbMJw1MZ8t5HZApcHrRKUc4W0kG879m7" crossorigin="anonymous">', unsafe_allow_html=True)
    st.markdown("""
                <button type="button" class="btn btn-primary">Primary</button>
                <button type="button" class="btn btn-secondary">Secondary</button>
                <button type="button" class="btn btn-success">Success</button>
                <button type="button" class="btn btn-danger">Danger</button>
                <button type="button" class="btn btn-warning">Warning</button>
                <button type="button" class="btn btn-info">Info</button>
                <button type="button" class="btn btn-light">Light</button>
                <button type="button" class="btn btn-dark">Dark</button>

                <button type="button" class="btn btn-link">Link</button>
                """, 
                unsafe_allow_html=True
                )
def card():
    st.markdown('<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.5/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-SgOJa3DmI69IUzQ2PVdRZhwQ+dy64/BUtbMJw1MZ8t5HZApcHrRKUc4W0kG879m7" crossorigin="anonymous">', unsafe_allow_html=True)
    st.markdown("""
                <div class="card" style="width: 18rem;">
                     <img src="https://ibhcxtnwnonsnycfgjay.supabase.co/storage/v1/object/sign/Meninos/Femeas/Betina.jpg?token=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1cmwiOiJNZW5pbm9zL0ZlbWVhcy9CZXRpbmEuanBnIiwiaWF0IjoxNzQ1Nzk2Njg1LCJleHAiOjE3NDY0MDE0ODV9.5SA9oA5cvr6dKMyZQBKCaWPNQL_DEJdS0yuf2P0cass" class="card-img-top" alt="...">
                     <div class="card-body">
                          <h5 class="card-title">Betina</h5>
                          <p class="card-text">Filha da Niltinha, femea castrada</p>
                          <a href="#" style="color:white;text-decoration:None" class="btn btn-primary">Ver perfil</a>
                     </div>
                </div>
                """, 
                unsafe_allow_html=True
                )

def atualizar(data, atividades):
    workbook = openpyxl.load_workbook("relatorio.xlsx")
    aba = workbook["Atividades"]
    aba.append([data, atividades])
    workbook.save("reLatorio.xlsx")
    
def nova_medicao(nro_ctr, nro_medicao, data_medicao, valor, nf, data_nf, data_pagto):
    w_medicao = openpyxl.load_workbook("relatorio.xlsx")
    aba1 = w_medicao["Medicao"]
    aba1.append([nro_ctr, nro_medicao, data_medicao, valor, nf, data_nf, data_pagto])
    w_medicao.save("relatorio.xlsx")


lista = ["INSERIR DIARIO DE OBRA", "INSERIR MEDICAO", "INSERIR CONTRATO", "CARD", "BOTAO", "CRIAR"]
a = st.sidebar.selectbox("Escolha a opção", lista)

if a==lista[0]:
 
    with st.form("Entrada"):
        data = st.date_input("Data:")
        atividades = st.text_area("Atividades")
        if st.form_submit_button("Confirmar"):
            atualizar(data, atividades)
        df1 = pd.read_excel("relatorio.xlsx", sheet_name=0)  
        st.dataframe(df1)  
    
elif a==lista[1]:
    with st.form("Medicao"):
        nro_ctr=st.text_input("numero do contrato")
        nro_medicao=st.text_input("Numero medicao")
        data_medicao = st.date_input("DAta da medicao")
        valor = st.text_input("valor da medicao")
        nf= st.text_input("Nota fiscal")
        data_nf= st.text_input("DAta da nota")
        data_pagto= st.text_input("Data pagto")
        if st.form_submit_button("Confirmar"):
            nova_medicao(nro_ctr, nro_medicao, data_medicao, valor, nf, data_nf, data_pagto)
    df = pd.read_excel("relatorio.xlsx", sheet_name=1) 
    st.dataframe(df) 
elif a=="CARD":
    card() 
elif a=="BOTAO":
    botao()
elif a=="CRIAR":
    lista_contratos_OBRAS = ["MASTER - 034-2022", "MASTER - 028-2023", "MASTER 019-2024", "CONSTRUTORA MENEGUETI - VG", "TECNBOMBAS - 007-2024", "TECNOBOMBAS - 004-2023", \
                         "SPARTACUS - 024-2024", "MILLENIUM - 009-2023", "MILLENIUM - 003-2024", "MILLENIUM 017-2024", "MILLENIUM - 008-2023", \
                         "COOMSER OBRA", "SPARTACUS - 013-2024", "SM7 - TANKS BR", "RST ENGENHARIA", "MARCIO SOUZA FARIAS",  "DIMBEL", "DA GARISTO"]
    ind = 0
    wb = openpyxl.Workbook()
    del wb['Sheet']
    
    #Cria as abas com os nomes dos contratos
    for nome in lista_contratos_OBRAS:
        
        ind = ind + 1
        print(ind)
        wb.create_sheet(title=f'{nome}', index=ind)
        
    wb.save("nome_teste.xlsx")
    wb1 = load_workbook("nome_teste.xlsx")
    
    #Abre as planilhas e coloca a linha de cabecalho Data e Atividades
    for nome in lista_contratos_OBRAS:
        wb1[nome].append(["Data","Atividades"])  
    wb1.save("nome_final.xlsx")    
    
        
    
 
        
